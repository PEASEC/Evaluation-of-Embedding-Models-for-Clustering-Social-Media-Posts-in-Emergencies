import numpy as np
from embeddings import get_document_embeddings
from openpyxl import load_workbook
import os
from tweet import Tweet
import clustering
import pandas as pd
from pandas.plotting import scatter_matrix
import seaborn as sns
import tensorflow as tf
import tensorflow_hub as hub
from sentence_transformers import SentenceTransformer
import matplotlib.pyplot as plt
from googletrans import Translator
import time

dirname = os.path.dirname(__file__)


models_types_to_use = [
    "word2vec_twitter",
    "word2vec_crisis_1",
    "word2vec_crisis_2",
    "glove",
    "fasttext_english",
    "fasttext_german",
    "sif",
    "use_base",  # discontinued
    "use_large",  # discontinued
    "infersent_glove",
    "infersent_fasttext",
    "sent2vec_unigrams",
    "sent2vec_bigrams",
    "nli_mean_base",  # discontinued
    "nli_mean_large",  # discontinued
    "nli_mean_sts_base",  # discontinued
    "nli_mean_sts_large",  # discontinued
]
models_to_use = {
    # The used SBERT and USE models are discontinued
    # Logic needs to be reimplemented if new models are used
}
CLUSTER_ALGORITHM = "KMeans"  # KMeans, SphericalKMeans, AgglomerativeClustering,
NORMALIZE = False
TRANSLATE = False
REMOVE_STOPWORDS = True
EXTENDED_DATA = False
APPEND_OTHER_FEATURES = False


def translate_tweets(tweets, dest="en"):

    for tweet in tweets:
        translator = Translator()
        src = tweet.language
        if src == "und":
            src = "auto"
        tweet.text = tweet.text.replace("---", "")
        try:
            tweet.text = translator.translate(tweet.text, src=src, dest=dest).text
        except Exception as e:
            print(tweet.text.encode("utf-8"))
            print(str(e))
            continue

    return tweets


def append_other_features(document_embeddings, tweets):
    for index, tweet in enumerate(tweets):
        document_embeddings[index] = np.append(
            document_embeddings[index],
            [  # tweet.location, tweet.time_distance, tweet.language
                int(tweet.num_retweets),
                len(tweet.text),
                int(tweet.author_distance),
                int(tweet.url),
                int(tweet.image),
            ],
        )

    document_embeddings = np.array(document_embeddings)

    return document_embeddings


def normalize_data(document_embeddings):

    document_embeddings /= np.linalg.norm(document_embeddings)
    # document_embeddings -= np.mean(document_embeddings)
    # document_embeddings /= np.std(document_embeddings)
    return document_embeddings


def get_nearest_tweets(document_embedding, cluster_centers, k, tweets):
    all_nearest_tweets = []

    for cluster_center in cluster_centers:
        nearest_tweets = []
        distances, indices = clustering.get_k_nearest(
            document_embedding, cluster_center, k
        )

        for index in indices:
            nearest_tweets.append(tweets[index])

        all_nearest_tweets.append(nearest_tweets)

        print("Cluster Center: ", cluster_center)
        for index, tweet in enumerate(nearest_tweets):
            print(
                str(index)
                + ". tweet text: "
                + tweet.text
                + " label: "
                + str(tweet.label)
            )

    return all_nearest_tweets


def plot_clustering(embeddings, labels):
    colors = len(labels) * ["violet"]
    for index, label in enumerate(labels):
        if label == 0:
            colors[index] = "red"
        elif label == 1:
            colors[index] = "blue"
        else:
            colors[index] = "green"

    embeddings = np.array(embeddings)
    embeddings = np.append(embeddings[:, 0:5], embeddings[:, 20:25], axis=1)
    data = pd.DataFrame(
        embeddings,
        columns=["x1", "x2", "x3", "x4", "x5" "x20", "x21", "x22", "x23", "x24", "x25"],
    )

    scatter_matrix(data, alpha=0.8, diagonal="hist", color=colors)
    plt.show()


def start_evaluation(model_type, model, document_type):
    output_file = open("output.txt", "a")
    output_file.write(
        "####################### " + model_type + " ####################### \n"
    )

    wb = load_workbook(os.path.join(dirname, "DATA/hochwasser.xlsx"))
    sheet = wb.active
    tweets = []
    tweet_texts = []

    if EXTENDED_DATA:
        for i in range(1, sheet.max_row + 1):
            tweet = Tweet(
                sheet["D" + str(i)].value[1:-1],
                sheet["B" + str(i)].value[1:-1],
                sheet["AA" + str(i)].value,
                sheet["F" + str(i)].value[1:-1],
                sheet["M" + str(i)].value[1:-1],
                sheet["H" + str(i)].value[1:-1],
                sheet["I" + str(i)].value[1:-1] != "",
                sheet["J" + str(i)].value[1:-1] != "",
                sheet["K" + str(i)].value[1:-1],
            )
            tweets.append(tweet)

        if TRANSLATE:
            tweets = translate_tweets(tweets, dest="en")

        tweet_texts = [tweet.text for tweet in tweets]
    else:
        for i in range(1, sheet.max_row + 1):
            tweet_texts.append(sheet["D" + str(i)].value[1:-1])

    start = time.time()
    document_embeddings, unknown_count, known_count = get_document_embeddings(
        tweet_texts, model_type, REMOVE_STOPWORDS, document_type, model
    )
    end = time.time()
    output_file.write("# Embedding Creation Time:")
    output_file.write(str(end - start))
    output_file.write("\n")

    print(
        "{} tokens were found as embeddings, {} tokens could not be found as embedding".format(
            known_count, unknown_count
        )
    )
    output_file.write("# Unknown Word Count:")
    output_file.write(str(unknown_count))
    output_file.write("\n")

    if APPEND_OTHER_FEATURES:
        document_embeddings = append_other_features(document_embeddings, tweets)

    if NORMALIZE:
        document_embeddings = normalize_data(document_embeddings)

    for k in range(2, 10):
        start = time.time()
        cluster_labels, cluster_centers = clustering.cluster(
            document_embeddings, k, CLUSTER_ALGORITHM
        )  # , cluster_centers
        end = time.time()
        output_file.write("k: " + str(k))
        output_file.write("# Cluster Creation Time:")
        output_file.write(str(end - start))
        output_file.write("\n")

        # for index, tweet in enumerate(tweets):
        #    tweet.set_cluster_label(cluster_labels[index])

        # nearest_neighbors = get_nearest_tweets(document_embeddings, cluster_centers, 5, tweets)

        silhouette, calinski_harabasz, davies_bouldin = clustering.evaluate(
            document_embeddings, labels=cluster_labels
        )

        output_file.write(
            "####################### Evaluation ####################### \n"
        )
        output_file.write("# Silhouette Score:")
        output_file.write(str(silhouette))
        output_file.write("\n")
        output_file.write("# Calinski Harabasz Score:")
        output_file.write(str(calinski_harabasz))
        output_file.write("\n")
        output_file.write("# Davies Bouldin Score:")
        output_file.write(str(davies_bouldin))
        output_file.write("\n\n\n")

    # plot_clustering(document_embeddings, cluster_labels)
    output_file.close()


def main():
    for model_type in models_types_to_use:
        if model_type in models_to_use:
            model = models_to_use[model_type]
        else:
            model = None

        start_evaluation(model_type, model, "Average")


if __name__ == "__main__":
    main()
